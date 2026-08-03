from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


class ExcelWriter:

    def __init__(self):
        pass

    def write(self, table_data, products, output_path):

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Supermarket Together"

        # -------------------------
        # Écriture des données
        # -------------------------

        for row_index, row in enumerate(table_data, start=1):

            for column_index, value in enumerate(row, start=1):

                worksheet.cell(
                    row=row_index,
                    column=column_index,
                    value=value
                )

        # -------------------------
        # Injection des traductions
        # -------------------------

        product_index = 0

        for row_index in range(2, worksheet.max_row + 1):

            product_cell = worksheet.cell(
                row=row_index,
                column=5
            )

        # On ignore les lignes qui ne sont pas des produits
            if not product_cell.value:
                continue

            if product_index >= len(products):
                break

            product = products[product_index]

        # Vérification que cette ligne correspond bien au produit
            original_product = f"{product.product_name} - {product.brand}"

            if product_cell.value.strip() != original_product.strip():
                continue

        # Produit traduit
            if product.translated_name:

                product_cell.value = (
                    f"{product.translated_name} - {product.brand}"
                )

            # Sinon on conserve automatiquement le texte anglais

            product_index += 1

        # -------------------------
        # Styles
        # -------------------------

        thin_border = Side(
            style="thin",
            color="BFBFBF"
        )

        border = Border(
            left=thin_border,
            right=thin_border,
            top=thin_border,
            bottom=thin_border
        )

        header_fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAD3"
        )

        # -------------------------
        # En-têtes
        # -------------------------

        for cell in worksheet[1]:

            cell.font = Font(
                bold=True
            )

            cell.fill = header_fill

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

            cell.border = border

        worksheet.row_dimensions[1].height = 45

        # -------------------------
        # Corps du tableau
        # -------------------------

        for row in worksheet.iter_rows(
            min_row=2,
            max_row=worksheet.max_row
        ):

            for cell in row:

                cell.alignment = Alignment(
                    vertical="center"
                )

                cell.border = border

        # -------------------------
        # Séparation des niveaux / licences
        # -------------------------

        level_border = Side(
            style="medium",
            color="808080"
        )

        license_border = Side(
            style="thick",
            color="404040"
        )

        for row_index in range(3, worksheet.max_row + 1):

            current_license = worksheet.cell(
              row=row_index,
              column=3
            ).value

            previous_license = worksheet.cell(
               row=row_index - 1,
               column=3
            ).value

            current_level = worksheet.cell(
              row=row_index,
              column=4
            ).value

            previous_level = worksheet.cell(
            row=row_index - 1,
            column=4
        ).value

            # Nouvelle licence/catégorie principale
            if current_license != previous_license:

                separator = license_border

            # Nouveau niveau dans la même licence
            elif current_level != previous_level:

                separator = level_border

            else:

                continue

            for cell in worksheet[row_index]:

                cell.border = Border(
                    left=cell.border.left,
                    right=cell.border.right,
                    top=separator,
                    bottom=cell.border.bottom
                )
                
        # -------------------------
        # Largeurs
        # -------------------------

        worksheet.column_dimensions["A"].width = 5
        worksheet.column_dimensions["B"].width = 6
        worksheet.column_dimensions["C"].width = 20
        worksheet.column_dimensions["D"].width = 6

        # PRODUCT
        worksheet.column_dimensions["E"].width = 35

        # DISP
        worksheet.column_dimensions["F"].width = 12

        # BOX / SMALL / B-S
        worksheet.column_dimensions["G"].width = 8
        worksheet.column_dimensions["H"].width = 9
        worksheet.column_dimensions["I"].width = 8

        worksheet.column_dimensions["J"].width = 3

        # LARGE / B-L
        worksheet.column_dimensions["K"].width = 9
        worksheet.column_dimensions["L"].width = 8

        worksheet.column_dimensions["M"].width = 3

        # BUY
        worksheet.column_dimensions["N"].width = 8

        worksheet.column_dimensions["O"].width = 3
        worksheet.column_dimensions["P"].width = 3
        worksheet.column_dimensions["Q"].width = 3

        # CATEGORIES
        worksheet.column_dimensions["R"].width = 22

        worksheet.column_dimensions["S"].width = 3

        # DISPLAYS
        worksheet.column_dimensions["T"].width = 32

        worksheet.column_dimensions["U"].width = 3
        worksheet.column_dimensions["V"].width = 3

        # -------------------------
        # Figer l'en-tête
        # -------------------------

        worksheet.freeze_panes = "A2"

        # -------------------------
        # Filtre automatique
        # -------------------------

        worksheet.auto_filter.ref = worksheet.dimensions

        workbook.save(output_path)